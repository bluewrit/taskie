"""Task and project CRUD endpoints (authenticated, assignment-aware)."""
import csv
import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session, joinedload

from .. import models
from ..database import UPLOAD_DIR, get_db
from ..deps import get_current_user
from ..models import FileAttachment, Project, ProjectMessage, Task, User
from ..schemas import (
    MessageIn,
    MessageOut,
    ProjectFileOut,
    ProjectIn,
    ProjectOut,
    TaskIn,
    TaskOut,
    TaskUpdate,
)

router = APIRouter(prefix="/api", tags=["tasks"])

VALID_STATUS = {"todo", "in_progress", "done", "blocked"}
VALID_PRIORITY = {"low", "medium", "high", "critical"}


def _validate(status: str | None, priority: str | None):
    if status and status not in VALID_STATUS:
        raise HTTPException(422, f"Invalid status '{status}'")
    if priority and priority not in VALID_PRIORITY:
        raise HTTPException(422, f"Invalid priority '{priority}'")


def _check_assignee(db: Session, assignee_id: int | None):
    if assignee_id and not db.get(User, assignee_id):
        raise HTTPException(422, "Assignee not found")


# ------------------------------------------------------------------ projects
@router.get("/projects", response_model=list[ProjectOut])
def list_projects(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(Project).order_by(Project.id).all()


@router.post("/projects", response_model=ProjectOut, status_code=201)
def create_project(payload: ProjectIn, db: Session = Depends(get_db),
                   _: User = Depends(get_current_user)):
    project = Project(**payload.model_dump())
    db.add(project)
    db.commit()
    db.refresh(project)
    return project


@router.delete("/projects/{project_id}", status_code=204)
def delete_project(project_id: int, db: Session = Depends(get_db),
                   _: User = Depends(get_current_user)):
    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")
    db.delete(project)
    db.commit()


# ------------------------------------------------------------------ project hub
def _get_project(db: Session, project_id: int) -> Project:
    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")
    return project


@router.get("/projects/{project_id}/files", response_model=list[ProjectFileOut])
def project_files(project_id: int, db: Session = Depends(get_db),
                  _: User = Depends(get_current_user)):
    """Shared files of a project: project-level uploads + every file
    attached to its tasks."""
    from sqlalchemy import or_

    _get_project(db, project_id)
    task_ids = [t.id for t in db.query(Task.id).filter(Task.project_id == project_id).all()]
    conditions = [FileAttachment.project_id == project_id]
    if task_ids:
        conditions.append(FileAttachment.task_id.in_(task_ids))
    query = (db.query(FileAttachment)
             .options(joinedload(FileAttachment.task))
             .filter(or_(*conditions)))
    out = []
    for f in query.order_by(FileAttachment.uploaded_at.desc(), FileAttachment.id.desc()).all():
        uploader = db.get(User, f.uploaded_by) if f.uploaded_by else None
        out.append(ProjectFileOut(
            id=f.id, filename=f.filename, mime_type=f.mime_type, extension=f.extension,
            size=f.size, uploaded_at=f.uploaded_at,
            task_id=f.task_id, task_title=f.task.title if f.task else None,
            uploaded_by_name=(uploader.full_name or uploader.username) if uploader else None,
        ))
    return out


@router.post("/projects/{project_id}/files", response_model=ProjectFileOut, status_code=201)
async def upload_project_file(project_id: int, file: UploadFile,
                              db: Session = Depends(get_db),
                              current: User = Depends(get_current_user)):
    """Upload a file shared with the whole project (not tied to a task)."""
    import uuid
    from pathlib import Path as _P

    _get_project(db, project_id)
    content = await file.read()
    if len(content) > 100 * 1024 * 1024:
        raise HTTPException(413, "File exceeds the 100 MB limit")
    if not content:
        raise HTTPException(422, "Empty file")
    filename = _P(file.filename or "upload.bin").name
    ext = _P(filename).suffix.lower()[:20]
    stored_name = f"{uuid.uuid4().hex}{ext}"
    (UPLOAD_DIR / stored_name).write_bytes(content)
    att = FileAttachment(
        task_id=None, project_id=project_id, filename=filename, stored_name=stored_name,
        mime_type=file.content_type or "application/octet-stream",
        extension=ext, size=len(content), uploaded_by=current.id,
    )
    db.add(att)
    db.commit()
    db.refresh(att)
    return ProjectFileOut(
        id=att.id, filename=att.filename, mime_type=att.mime_type, extension=att.extension,
        size=att.size, uploaded_at=att.uploaded_at,
        uploaded_by_name=current.full_name or current.username,
    )


@router.get("/projects/{project_id}/messages", response_model=list[MessageOut])
def project_messages(project_id: int, after: int = 0, db: Session = Depends(get_db),
                     _: User = Depends(get_current_user)):
    _get_project(db, project_id)
    query = (db.query(ProjectMessage)
             .options(joinedload(ProjectMessage.user))
             .filter(ProjectMessage.project_id == project_id, ProjectMessage.id > after)
             .order_by(ProjectMessage.id))
    return query.all()


@router.post("/projects/{project_id}/messages", response_model=MessageOut, status_code=201)
def post_project_message(project_id: int, payload: MessageIn,
                         db: Session = Depends(get_db),
                         current: User = Depends(get_current_user)):
    _get_project(db, project_id)
    msg = ProjectMessage(project_id=project_id, user_id=current.id, body=payload.body.strip()[:4000])
    db.add(msg)
    db.commit()
    db.refresh(msg)
    return msg


# ------------------------------------------------------------------ tasks
@router.get("/tasks", response_model=list[TaskOut])
def list_tasks(status: str | None = None, project_id: int | None = None,
               assignee_id: int | None = None, mine: bool = False,
               unassigned: bool = False, q: str | None = None,
               db: Session = Depends(get_db),
               current: User = Depends(get_current_user)):
    query = (db.query(Task)
             .options(joinedload(Task.files), joinedload(Task.project), joinedload(Task.assignee)))
    if status:
        query = query.filter(Task.status == status)
    if project_id:
        query = query.filter(Task.project_id == project_id)
    if mine:
        query = query.filter(Task.assignee_id == current.id)
    if unassigned:
        query = query.filter(Task.assignee_id.is_(None))
    elif assignee_id:
        query = query.filter(Task.assignee_id == assignee_id)
    if q:
        like = f"%{q}%"
        query = query.filter(Task.title.ilike(like) | Task.description.ilike(like))
    return query.order_by(Task.agent_score.desc(), Task.id.desc()).all()


@router.post("/tasks", response_model=TaskOut, status_code=201)
def create_task(payload: TaskIn, db: Session = Depends(get_db),
                current: User = Depends(get_current_user)):
    _validate(payload.status, payload.priority)
    if payload.project_id and not db.get(Project, payload.project_id):
        raise HTTPException(422, "Project not found")
    _check_assignee(db, payload.assignee_id)
    task = Task(**payload.model_dump())
    db.add(task)
    db.commit()
    db.refresh(task)
    return task


# ------------------------------------------------------------------ import
_STATUS_ALIASES = {
    "todo": "todo", "to do": "todo", "to-do": "todo", "open": "todo", "new": "todo",
    "backlog": "todo", "not started": "todo",
    "in progress": "in_progress", "in_progress": "in_progress", "inprogress": "in_progress",
    "doing": "in_progress", "started": "in_progress", "wip": "in_progress",
    "done": "done", "complete": "done", "completed": "done", "finished": "done", "closed": "done",
    "blocked": "blocked", "on hold": "blocked", "stuck": "blocked",
}
_PRIORITY_ALIASES = {
    "low": "low", "minor": "low",
    "medium": "medium", "med": "medium", "normal": "medium", "moderate": "medium",
    "high": "high", "major": "high", "urgent": "high",
    "critical": "critical", "crit": "critical", "blocker": "critical", "p0": "critical",
}
_COLUMN_ALIASES = {
    "title": ["title", "task", "taskname", "name", "summary", "subject"],
    "description": ["description", "desc", "details", "notes", "comments"],
    "status": ["status", "state", "stage"],
    "priority": ["priority", "prio", "severity", "importance"],
    "due_date": ["duedate", "due", "deadline", "date", "dueby", "targetdate"],
    "assignee": ["assignee", "assignedto", "owner", "assigned", "user", "username",
                 "member", "responsible"],
    "project": ["project", "projectname", "initiative", "workstream"],
    "estimated_minutes": ["estimatedminutes", "estimate", "estimateminutes", "minutes",
                          "effort", "time", "estimation"],
    "hours": ["hours", "estimatedhours", "esthours"],
    "progress": ["progress", "percent", "percentcomplete", "complete", "completion", "done"],
    "tags": ["tags", "labels", "tag"],
}
_PROJECT_COLORS = ["#0a84ff", "#30d158", "#ff9f0a", "#bf5af2", "#64d2ff", "#ff375f",
                   "#ffd60a", "#ac8e68"]


def _norm_header(value) -> str:
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


def _parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y",
                "%b %d %Y", "%B %d, %Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"unrecognised date '{s}'")


def _rows_from_upload(filename: str, data: bytes):
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        try:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:  # corrupt / password-protected / fake extension
            raise HTTPException(422, f"Could not read Excel file: {exc}")
        ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]
    if name.endswith(".csv"):
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = data.decode("latin-1")
        return [list(r) for r in csv.reader(io.StringIO(text))]
    raise HTTPException(422, "Unsupported file type — upload an .xlsx or .csv file "
                             "(for .xls, re-save as .xlsx first)")


@router.post("/tasks/import", status_code=201)
async def import_tasks(file: UploadFile = File(...), db: Session = Depends(get_db),
                       _: User = Depends(get_current_user)):
    """Create tasks in bulk from an Excel (.xlsx) or CSV file.

    The first non-empty row must be a header row; columns are matched
    flexibly (e.g. 'Task', 'Title', 'Summary' all work). Unknown projects
    are created on the fly; unknown assignees leave the task unassigned.
    """
    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(422, "File too large (10 MB max)")
    rows = _rows_from_upload(file.filename, data)
    rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
    if not rows:
        raise HTTPException(422, "File contains no data")
    if len(rows) > 1001:
        raise HTTPException(422, "Too many rows (1000 max per import)")

    header = [_norm_header(c) for c in rows[0]]
    colmap = {}  # canonical -> index
    for canon, aliases in _COLUMN_ALIASES.items():
        for idx, h in enumerate(header):
            if h in aliases and canon not in colmap:
                colmap[canon] = idx
                break
    if "title" not in colmap:
        raise HTTPException(422, "No title column found — the header row needs a "
                                 "'Title' (or 'Task'/'Name'/'Summary') column")

    users = db.query(User).all()
    by_username = {u.username.lower(): u for u in users}
    by_name = {(u.full_name or "").lower(): u for u in users if u.full_name}
    projects = {p.name.lower(): p for p in db.query(Project).all()}

    created, skipped, projects_created = 0, [], []

    def cell(row, canon):
        idx = colmap.get(canon)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return v

    for row_no, row in enumerate(rows[1:], start=2):
        try:
            title = str(cell(row, "title") or "").strip()
            if not title:
                skipped.append({"row": row_no, "reason": "missing title"})
                continue

            status_raw = str(cell(row, "status") or "").strip().lower()
            status = _STATUS_ALIASES.get(status_raw, "todo" if not status_raw else None)
            if status is None:
                skipped.append({"row": row_no, "reason": f"unknown status '{status_raw}'"})
                continue

            prio_raw = str(cell(row, "priority") or "").strip().lower()
            priority = _PRIORITY_ALIASES.get(prio_raw, "medium" if not prio_raw else None)
            if priority is None:
                skipped.append({"row": row_no, "reason": f"unknown priority '{prio_raw}'"})
                continue

            due = None
            try:
                due = _parse_date(cell(row, "due_date"))
            except ValueError as exc:
                skipped.append({"row": row_no, "reason": str(exc)})
                continue

            assignee = None
            raw_assignee = str(cell(row, "assignee") or "").strip()
            if raw_assignee:
                assignee = by_username.get(raw_assignee.lower()) or by_name.get(raw_assignee.lower())

            project = None
            raw_project = str(cell(row, "project") or "").strip()
            if raw_project:
                project = projects.get(raw_project.lower())
                if project is None:
                    project = Project(name=raw_project,
                                      color=_PROJECT_COLORS[len(projects) % len(_PROJECT_COLORS)])
                    db.add(project)
                    db.flush()
                    projects[raw_project.lower()] = project
                    projects_created.append(raw_project)

            est = 60
            if "hours" in colmap and cell(row, "hours") not in (None, ""):
                try:
                    est = max(15, int(float(cell(row, "hours")) * 60))
                except (TypeError, ValueError):
                    pass
            elif cell(row, "estimated_minutes") not in (None, ""):
                try:
                    est = max(15, int(float(cell(row, "estimated_minutes"))))
                except (TypeError, ValueError):
                    pass

            progress = 100 if status == "done" else 0
            if status != "done" and cell(row, "progress") not in (None, ""):
                try:
                    progress = max(0, min(100, int(float(cell(row, "progress")))))
                except (TypeError, ValueError):
                    pass

            db.add(Task(
                title=title[:300],
                description=str(cell(row, "description") or "").strip(),
                status=status,
                priority=priority,
                due_date=due,
                assignee_id=assignee.id if assignee else None,
                project_id=project.id if project else None,
                estimated_minutes=est,
                progress=progress,
                tags=str(cell(row, "tags") or "").strip(),
            ))
            created += 1
        except HTTPException:
            raise
        except Exception as exc:  # defensive: never 500 on one bad row
            skipped.append({"row": row_no, "reason": str(exc)[:120]})

    db.commit()
    return {"created": created, "skipped": skipped, "projects_created": projects_created,
            "columns_found": sorted(colmap.keys())}


@router.get("/tasks/{task_id}", response_model=TaskOut)
def get_task(task_id: int, db: Session = Depends(get_db),
             _: User = Depends(get_current_user)):
    task = db.get(Task, task_id)
    if not task:
        raise HTTPException(404, "Task not found")
    return task


@router.put("/tasks/{task_id}", response_model=TaskOut)
def update_task(task_id: int, payload: TaskUpdate, db: Session = Depends(get_db),
                _: User = Depends(get_current_user)):
    task = db.get(Task, task_id)
    if not task:
        raise HTTPException(404, "Task not found")
    data = payload.model_dump(exclude_unset=True)
    _validate(data.get("status"), data.get("priority"))
    if "project_id" in data and data["project_id"] and not db.get(Project, data["project_id"]):
        raise HTTPException(422, "Project not found")
    if "assignee_id" in data:
        _check_assignee(db, data["assignee_id"])
    was_done = task.status == "done"
    for key, value in data.items():
        setattr(task, key, value)
    if task.status == "done" and not was_done:
        task.completed_at = datetime.utcnow()
        task.progress = 100
    if task.status != "done" and was_done:
        task.completed_at = None
    task.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(task)
    return task


@router.post("/tasks/{task_id}/complete", response_model=TaskOut)
def complete_task(task_id: int, db: Session = Depends(get_db),
                  _: User = Depends(get_current_user)):
    task = db.get(Task, task_id)
    if not task:
        raise HTTPException(404, "Task not found")
    task.status = "done"
    task.progress = 100
    task.completed_at = datetime.utcnow()
    task.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(task)
    return task


@router.delete("/tasks/{task_id}", status_code=204)
def delete_task(task_id: int, db: Session = Depends(get_db),
                _: User = Depends(get_current_user)):
    task = db.get(Task, task_id)
    if not task:
        raise HTTPException(404, "Task not found")
    db.delete(task)
    db.commit()
